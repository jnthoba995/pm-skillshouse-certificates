import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

INPUT_JSON = "document.json"
OUTPUT_XLSX = "docai_table_output.xlsx"

def load_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))

def get_text_anchor_text(doc_text, text_anchor):
    if not text_anchor:
        return ""
    parts = []
    for seg in text_anchor.get("textSegments", []):
        start = int(seg.get("startIndex", 0))
        end = int(seg.get("endIndex", 0))
        parts.append(doc_text[start:end])
    return "".join(parts)

def layout_text(doc_text, layout):
    if not layout:
        return ""
    return normalize_ws(get_text_anchor_text(doc_text, layout.get("textAnchor")))

def normalize_ws(text):
    return re.sub(r"\s+", " ", (text or "")).strip()

def clean_name(text):
    text = normalize_ws(text)
    text = re.sub(r"[^A-Za-z\s\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return ""
    parts = [p for p in text.split() if len(p) >= 2]
    return " ".join(p.capitalize() for p in parts)

def clean_id(text):
    digits = re.sub(r"\D", "", text or "")
    if len(digits) >= 13:
        return digits[:13]
    if len(digits) >= 10:
        return digits
    return ""

def clean_phone(text):
    digits = re.sub(r"\D", "", text or "")
    if len(digits) >= 10:
        return digits[:10]
    return ""

def clean_email(text):
    text = normalize_ws(text).lower()
    m = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", text)
    return m.group(0) if m else ""

def extract_row_cells(doc_text, row_obj):
    cells = row_obj.get("cells", [])
    return [layout_text(doc_text, c.get("layout")) for c in cells]

def infer_headers(header_rows):
    if not header_rows:
        return []
    merged = []
    max_len = max(len(r) for r in header_rows)
    for i in range(max_len):
        bits = []
        for row in header_rows:
            if i < len(row) and row[i]:
                bits.append(row[i])
        merged.append(normalize_ws(" ".join(bits)))
    return merged

def map_header(header):
    h = header.lower()
    if "name" in h and "surname" not in h:
        return "name"
    if "surname" in h:
        return "surname"
    if "id" in h or "date of birth" in h:
        return "id_number"
    if "contact" in h:
        return "phone"
    if "email" in h:
        return "email"
    if "gender" in h:
        return "gender"
    if "employment" in h:
        return "employment"
    if "credit" in h:
        return "credit_report"
    if "financial" in h or "advice" in h:
        return "financial_advice"
    if "age" in h:
        return "age"
    return None

def score_row(row):
    score = 0
    if row["name"]:
        score += 1
    if row["surname"]:
        score += 1
    if len(row["id_number"]) == 13:
        score += 2
    elif row["id_number"]:
        score += 1
    if len(row["phone"]) == 10:
        score += 1
    if row["email"]:
        score += 1
    if score >= 4:
        return "high"
    if score >= 2:
        return "medium"
    return "low"

def row_has_signal(row):
    return any([
        row["name"],
        row["surname"],
        row["id_number"],
        row["phone"],
        row["email"],
    ])

def dedupe(rows):
    seen = set()
    out = []
    for row in rows:
        key = (
            row["name"],
            row["surname"],
            row["id_number"],
            row["phone"],
            row["email"],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out

def main():
    doc = load_json(INPUT_JSON)
    doc_text = doc.get("text", "")
    pages = doc.get("pages", [])

    extracted = []

    for page in pages:
        for table in page.get("tables", []):
            header_rows = [extract_row_cells(doc_text, r) for r in table.get("headerRows", [])]
            body_rows = [extract_row_cells(doc_text, r) for r in table.get("bodyRows", [])]

            headers = infer_headers(header_rows)
            mapped = [map_header(h) for h in headers]

            for body in body_rows:
                row = {
                    "name": "",
                    "surname": "",
                    "id_number": "",
                    "phone": "",
                    "email": "",
                    "gender": "",
                    "employment": "",
                    "credit_report": "",
                    "financial_advice": "",
                    "age": "",
                    "raw": " | ".join(body),
                }

                for idx, cell in enumerate(body):
                    key = mapped[idx] if idx < len(mapped) else None
                    if key == "name":
                        row["name"] = clean_name(cell)
                    elif key == "surname":
                        row["surname"] = clean_name(cell)
                    elif key == "id_number":
                        row["id_number"] = clean_id(cell)
                    elif key == "phone":
                        row["phone"] = clean_phone(cell)
                    elif key == "email":
                        row["email"] = clean_email(cell)
                    elif key == "gender":
                        row["gender"] = normalize_ws(cell)
                    elif key == "employment":
                        row["employment"] = normalize_ws(cell)
                    elif key == "credit_report":
                        row["credit_report"] = normalize_ws(cell)
                    elif key == "financial_advice":
                        row["financial_advice"] = normalize_ws(cell)
                    elif key == "age":
                        row["age"] = normalize_ws(cell)

                if row_has_signal(row):
                    row["confidence"] = score_row(row)
                    extracted.append(row)

    extracted = dedupe(extracted)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    headers = [
        "Row", "Name", "Surname", "ID Number", "Phone", "Email",
        "Gender", "Age", "Employment", "Credit Report",
        "Financial Advice", "Confidence", "Raw Source"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, row in enumerate(extracted, start=1):
        ws.append([
            i,
            row["name"],
            row["surname"],
            row["id_number"],
            row["phone"],
            row["email"],
            row["gender"],
            row["age"],
            row["employment"],
            row["credit_report"],
            row["financial_advice"],
            row["confidence"],
            row["raw"],
        ])

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total extracted rows"
    summary["B1"] = len(extracted)
    summary["A3"] = "High confidence"
    summary["B3"] = sum(1 for r in extracted if r["confidence"] == "high")
    summary["A4"] = "Medium confidence"
    summary["B4"] = sum(1 for r in extracted if r["confidence"] == "medium")
    summary["A5"] = "Low confidence"
    summary["B5"] = sum(1 for r in extracted if r["confidence"] == "low")

    for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws.column_dimensions[col].width = 22
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14

    wb.save(OUTPUT_XLSX)
    print(f"OK: wrote {OUTPUT_XLSX} with {len(extracted)} rows")

if __name__ == "__main__":
    main()
