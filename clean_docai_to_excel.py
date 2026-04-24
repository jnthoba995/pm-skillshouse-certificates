import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

INPUT_JSON = "document.json"
OUTPUT_XLSX = "docai_clean_output.xlsx"

def norm_text(v: str) -> str:
    return re.sub(r"\s+", " ", (v or "")).strip()

def clean_name(v: str) -> str:
    v = norm_text(v)
    v = re.sub(r"^[\W\d_xX+*#'\"`.,:;><=/-]+", "", v)
    v = re.sub(r"[\W_]+$", "", v)
    v = re.sub(r"[^A-Za-z\s\-]", " ", v)
    v = re.sub(r"\s+", " ", v).strip()
    if not v:
        return ""
    parts = [p for p in v.split() if len(p) >= 2]
    return " ".join(p.capitalize() for p in parts)

def clean_id(v: str) -> str:
    digits = re.sub(r"\D", "", v or "")
    if len(digits) >= 13:
        return digits[:13]
    if len(digits) >= 10:
        return digits
    return ""

def clean_phone(v: str) -> str:
    digits = re.sub(r"\D", "", v or "")
    if len(digits) >= 10:
        return digits[:10]
    return ""

def clean_email(v: str) -> str:
    v = norm_text(v).lower()
    m = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", v)
    return m.group(0) if m else ""

def score_row(name, surname, id_number, phone, email):
    score = 0
    if name: score += 1
    if surname: score += 1
    if len(id_number) == 13: score += 2
    elif id_number: score += 1
    if len(phone) == 10: score += 1
    if email: score += 1
    if score >= 4:
        return "high"
    if score >= 2:
        return "medium"
    return "low"

def extract_lines(doc_text: str):
    lines = [norm_text(x) for x in doc_text.splitlines()]
    return [x for x in lines if x]

def looks_like_noise(line: str) -> bool:
    bad_patterns = [
        r"^mind my money",
        r"^attendance register",
        r"^name of training provider",
        r"^facilitator",
        r"^session language",
        r"^province",
        r"^municipality",
        r"^register code",
        r"^employment status",
        r"^do you",
        r"^email address",
        r"^signature$",
        r"^liberty collects",
        r"^www\.mindmymoney",
        r"^standard bank",
        r"^liberty$",
    ]
    low = line.lower()
    return any(re.search(p, low) for p in bad_patterns)

def row_candidate(line: str):
    raw = line

    email = clean_email(raw)
    phone = clean_phone(raw)
    id_number = clean_id(raw)

    working = raw
    if email:
        working = working.replace(email, " ")
    if phone:
        working = working.replace(phone, " ")
    if id_number:
        working = working.replace(id_number, " ")

    working = re.sub(r"\b(?:yes|no|employed|unemployed|urban|rural|male|female|m|f|x|w|c|i|a)\b", " ", working, flags=re.I)
    working = re.sub(r"[\d]+", " ", working)
    working = re.sub(r"[^A-Za-z\s\-]", " ", working)
    working = re.sub(r"\s+", " ", working).strip()

    parts = [p for p in working.split() if len(p) >= 2]

    name = ""
    surname = ""

    if len(parts) >= 2:
        name = clean_name(parts[0])
        surname = clean_name(" ".join(parts[1:3]))
    elif len(parts) == 1:
        name = clean_name(parts[0])

    confidence = score_row(name, surname, id_number, phone, email)

    return {
        "raw": raw,
        "name": name,
        "surname": surname,
        "id_number": id_number,
        "phone": phone,
        "email": email,
        "confidence": confidence,
    }

def is_useful_row(row):
    filled = sum(bool(row[k]) for k in ["name", "surname", "id_number", "phone", "email"])
    if filled < 2:
        return False
    if not any([row["id_number"], row["phone"], row["email"]]):
        return False
    return True

def dedupe(rows):
    seen = set()
    out = []
    for r in rows:
        key = (r["name"], r["surname"], r["id_number"], r["phone"], r["email"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out

def main():
    path = Path(INPUT_JSON)
    if not path.exists():
        raise FileNotFoundError(f"Missing {INPUT_JSON}")

    data = json.loads(path.read_text(encoding="utf-8"))
    text = data.get("text", "")
    lines = extract_lines(text)

    candidates = []
    for line in lines:
        if looks_like_noise(line):
            continue
        row = row_candidate(line)
        if is_useful_row(row):
            candidates.append(row)

    rows = dedupe(candidates)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    headers = ["Row", "Name", "Surname", "ID Number", "Phone", "Email", "Confidence", "Raw Source"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row["name"],
            row["surname"],
            row["id_number"],
            row["phone"],
            row["email"],
            row["confidence"],
            row["raw"],
        ])

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total extracted rows"
    summary["B1"] = len(rows)
    summary["A3"] = "High confidence"
    summary["B3"] = sum(1 for r in rows if r["confidence"] == "high")
    summary["A4"] = "Medium confidence"
    summary["B4"] = sum(1 for r in rows if r["confidence"] == "medium")
    summary["A5"] = "Low confidence"
    summary["B5"] = sum(1 for r in rows if r["confidence"] == "low")

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 22
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14

    wb.save(OUTPUT_XLSX)
    print(f"OK: wrote {OUTPUT_XLSX} with {len(rows)} rows")

if __name__ == "__main__":
    main()
